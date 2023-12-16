import pandas as pd
from openpyxl import load_workbook

def read_planning_data(file_path):
    planning_df = pd.read_excel(file_path, sheet_name='ANNEE 2023')
    pointage_df = pd.read_excel(file_path, sheet_name='POINTAGE')
    return planning_df, pointage_df

def get_of_code_mappings(pointage_df):
    of_code_mappings = {}
    for _, row in pointage_df.iterrows():
        project_code = row['Unnamed: 1']
        if pd.notna(project_code):
            project_code = ''.join(project_code.split())
            of_code_mappings[project_code] = {
                'INTERNE': row['Unnamed: 2'],
                'EXTERNE': row['Unnamed: 3']
            }
    return of_code_mappings

def extract_week_data(planning_df, week_number, of_code_mappings):
    start_col_index = 273 + (week_number - 39) * 7
    end_col_index = start_col_index + 6
    samuel_murez_row_range = (546, 550)

    week_data = planning_df.iloc[samuel_murez_row_range[0]:samuel_murez_row_range[1], start_col_index:end_col_index]

    project_info = {}
    for col in week_data.columns:
        for row in week_data[col].dropna().unique():
            if ' ' in row:
                project_code, customer_name = row.split(' ', 1)
                of_codes = of_code_mappings.get(project_code, {'INTERNE': None, 'EXTERNE': None})
                project_info[project_code] = {
                    'Customer Name': customer_name,
                    'OF Codes': of_codes
                }
    return project_info

def prefill_of_numbers_with_openpyxl(file_path, sheet_name, project_info, start_row_index):
    workbook = load_workbook(file_path)
    worksheet = workbook[sheet_name]

    for project_code, details in project_info.items():
        response = input(f"Is the project code {project_code} internal or external? (i/e): ").strip().lower()
        of_code = details['OF Codes']['INTERNE'] if response == 'i' else details['OF Codes']['EXTERNE']

        current_row = start_row_index + 1
        while worksheet[f'J{current_row}'].value is not None and current_row <= worksheet.max_row:
            current_row += 1

        if current_row <= worksheet.max_row:
            # Fill OF code
            worksheet[f'J{current_row}'].value = of_code
            # Fill codes based on internal or external
            if response == 'i':
                worksheet[f'K{current_row}'].value = 'BAIS'  # Column to the right (+1)
                worksheet[f'L{current_row}'].value = 'BAI'   # Column +2
            elif response == 'e':
                # First entry
                worksheet[f'K{current_row}'].value = 'MVVS'  # Column to the right (+1)
                worksheet[f'L{current_row}'].value = 'MVV'   # Column +2
                # Second entry (if needed)
                current_row += 1
                if current_row <= worksheet.max_row:
                    worksheet[f'J{current_row}'].value = of_code
                    worksheet[f'K{current_row}'].value = 'MSES'  # Column to the right (+1)
                    worksheet[f'L{current_row}'].value = 'MSE'   # Column +2

    return workbook

def main():
    planning_file_path = 'Planning SAV.xlsx'
    timesheet_file_path = 'Feuille_pointage_.xlsx'
    sheet_name = 'S(45)'

    planning_df, pointage_df = read_planning_data(planning_file_path)
    of_code_mappings = get_of_code_mappings(pointage_df)

    week_number = int(input("Enter the week number: "))
    project_info = extract_week_data(planning_df, week_number, of_code_mappings)

    updated_workbook = prefill_of_numbers_with_openpyxl(timesheet_file_path, sheet_name, project_info, 10)
    updated_timesheet_path = 'updated_timesheet.xlsx'
    updated_workbook.save(updated_timesheet_path)
    print(f"Updated timesheet saved to '{updated_timesheet_path}'.")

if __name__ == "__main__":
    main()